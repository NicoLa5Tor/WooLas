from io import BytesIO
import re
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook

from schemas.product import PreviewRow, UpdateSummary


ALLOWED_FIELDS = {
    "name",
    "slug",
    "type",
    "status",
    "featured",
    "catalog_visibility",
    "regular_price",
    "sale_price",
    "date_on_sale_from",
    "date_on_sale_to",
    "description",
    "short_description",
    "manage_stock",
    "stock_quantity",
    "stock_status",
    "backorders",
    "sold_individually",
    "low_stock_amount",
    "weight",
    "length",
    "width",
    "height",
    "shipping_class",
    "virtual",
    "downloadable",
    "images",
    "image_principal_id",
    "image_galeria_ids",
    "image_nombres",
    "attr1_nombre",
    "attr1_valores",
    "attr2_nombre",
    "attr2_valores",
    "category_ids",
    "tag_ids",
    "upsell_ids",
    "cross_sell_ids",
    "meta_key",
    "meta_value",
}

HEADER_ALIASES = {
    "ID del producto": "product_id",
    "SKU": "sku",
    "Nombre": "name",
    "Slug": "slug",
    "Tipo": "type",
    "Estado": "status",
    "Destacado": "featured",
    "Visibilidad catalogo": "catalog_visibility",
    "Precio regular": "regular_price",
    "Precio oferta": "sale_price",
    "Oferta desde": "date_on_sale_from",
    "Oferta hasta": "date_on_sale_to",
    "Descripcion completa": "description",
    "Descripcion corta": "short_description",
    "Gestionar stock": "manage_stock",
    "Cantidad stock": "stock_quantity",
    "Estado stock": "stock_status",
    "Backorders": "backorders",
    "Venta individual": "sold_individually",
    "Alerta stock bajo": "low_stock_amount",
    "Peso": "weight",
    "Largo": "length",
    "Ancho": "width",
    "Alto": "height",
    "Clase de envio": "shipping_class",
    "Virtual": "virtual",
    "Descargable": "downloadable",
    "ID imagen principal": "image_principal_id",
    "IDs galeria": "image_galeria_ids",
    "Nombres imagenes": "image_nombres",
    "Atributo 1 nombre": "attr1_nombre",
    "Atributo 1 valores": "attr1_valores",
    "Atributo 2 nombre": "attr2_nombre",
    "Atributo 2 valores": "attr2_valores",
    "IDs categorias": "category_ids",
    "IDs etiquetas": "tag_ids",
    "IDs upsells": "upsell_ids",
    "IDs cross-sells": "cross_sell_ids",
    "Meta key": "meta_key",
    "Meta value": "meta_value",
}
TEMPLATE_SAMPLE_MARKERS = {
    "product_id": "12345",
    "sku": "03330154",
    "name": "Producto ejemplo",
}

FIELD_GROUPS = {
    "images": ["image_principal_id", "image_galeria_ids", "image_nombres"],
}

# Fields where empty cell means "clear field" instead of "skip".
# Boolean/numeric/enum fields stay out (blank there is ambiguous).
BLANKABLE_FIELDS = {
    "name", "slug", "sku",
    "regular_price", "sale_price", "date_on_sale_from", "date_on_sale_to",
    "description", "short_description",
    "weight", "length", "width", "height", "shipping_class",
    "catalog_visibility",
    "meta_key", "meta_value",
    "category_ids", "tag_ids", "upsell_ids", "cross_sell_ids",
    "image_principal_id", "image_galeria_ids", "image_nombres", "images",
    "attr1_nombre", "attr1_valores", "attr2_nombre", "attr2_valores",
}


def _normalize_header(header: object) -> str:
    value = "" if header is None else str(header).strip()
    if value.startswith("* "):
        value = value[2:].strip()
    return HEADER_ALIASES.get(value, value)


def _header_row_index(rows: list[tuple[Any, ...] | tuple]) -> int:
    return 2 if len(rows) >= 3 and any(str(value).strip() for value in rows[2] if value is not None) else 0


def parse_excel(file_bytes: bytes, *, normalize_headers: bool = True) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_row_index = _header_row_index(rows)
    headers = [(_normalize_header(value) if normalize_headers else ("" if value is None else str(value).strip())) for value in rows[header_row_index]]
    parsed_rows: list[dict[str, str]] = []

    for row in rows[header_row_index + 1 :]:
        parsed_row: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            parsed_row[header] = "" if value is None else str(value)
        if all(parsed_row.get(key, "") == value for key, value in TEMPLATE_SAMPLE_MARKERS.items()):
            continue
        if any(value != "" for value in parsed_row.values()):
            parsed_rows.append(parsed_row)

    return headers, parsed_rows


def validate_mapping(headers: list[str], id_column: str, value_column: str, wc_field: str) -> None:
    if wc_field not in ALLOWED_FIELDS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid WooCommerce field")
    if id_column not in headers or value_column not in headers:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selected columns were not found in the file")


def validate_selected_fields(headers: list[str], selected_fields: list[str]) -> None:
    if not selected_fields:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes seleccionar al menos una columna para editar")
    invalid_fields = [field for field in selected_fields if field not in ALLOWED_FIELDS]
    if invalid_fields:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid WooCommerce fields: {', '.join(invalid_fields)}")
    missing_columns: list[str] = []
    for field in selected_fields:
        required_headers = FIELD_GROUPS.get(field, [field])
        for header in required_headers:
            if header not in headers and header not in missing_columns:
                missing_columns.append(header)
    if missing_columns:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Selected columns were not found in the file: {', '.join(missing_columns)}")


def normalize_field_value(field: str, value: str):
    if field in {"stock_quantity", "low_stock_amount", "download_limit", "download_expiry"}:
        return int(float(value)) if value != "" else 0
    if field in {"regular_price", "sale_price"}:
        return str(value)
    if field == "status":
        return "draft" if str(value).strip().lower() in {"draft", "borrador", "inactive"} else "publish"
    if field in {"featured", "manage_stock", "sold_individually", "virtual", "downloadable"}:
        return str(value).strip().lower() in {"1", "true", "yes", "si", "sí", "y"}
    if field in {"description", "short_description"}:
        paragraphs = [line.strip() for line in str(value).replace("\r\n", "\n").split("\n") if line.strip()]
        return "".join(f"<p>{paragraph}</p>" for paragraph in paragraphs) if paragraphs else ""
    return value


def _plain_text(value: str) -> str:
    text = str(value or "")
    text = re.sub(r"</p>\s*<p>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def _split_values(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;,\n]+", str(value)) if item.strip()]


def _split_int_values(value: str) -> list[int]:
    parsed: list[int] = []
    for item in _split_values(value):
        try:
            parsed.append(int(float(item)))
        except ValueError as exc:
            raise ValueError(f"Valor numerico invalido: {item}") from exc
    return parsed


def _expand_selected_fields(selected_fields: list[str]) -> list[str]:
    expanded: list[str] = []
    for field in selected_fields:
        for expanded_field in FIELD_GROUPS.get(field, [field]):
            if expanded_field not in expanded:
                expanded.append(expanded_field)
    return expanded


def _join_product_ids(items: list[dict[str, Any]] | None) -> str:
    return "; ".join(str(item.get("id")) for item in (items or []) if item.get("id") is not None)


def _join_image_ids(items: list[dict[str, Any]] | None) -> str:
    return "; ".join(str(item.get("id")) for item in (items or []) if item.get("id") is not None)


def _join_image_names(items: list[dict[str, Any]] | None) -> str:
    names: list[str] = []
    for item in items or []:
        value = str(item.get("name") or item.get("src") or "").strip()
        if value:
            names.append(value)
    return "; ".join(names)


def _join_attribute_values(attribute: dict[str, Any] | None) -> str:
    return "; ".join(str(option).strip() for option in (attribute or {}).get("options", []) if str(option).strip())


def _find_attribute(product: dict[str, Any], index: int) -> dict[str, Any] | None:
    attributes = product.get("attributes") or []
    return attributes[index] if len(attributes) > index else None


def _current_field_value(product: dict[str, Any], field: str) -> str:
    dimensions = product.get("dimensions") or {}
    first_attribute = _find_attribute(product, 0)
    second_attribute = _find_attribute(product, 1)
    mapping: dict[str, str] = {
        "name": str(product.get("name") or ""),
        "slug": str(product.get("slug") or ""),
        "sku": str(product.get("sku") or ""),
        "type": str(product.get("type") or ""),
        "status": str(product.get("status") or ""),
        "featured": "true" if product.get("featured") else "false",
        "catalog_visibility": str(product.get("catalog_visibility") or ""),
        "regular_price": str(product.get("regular_price") or ""),
        "sale_price": str(product.get("sale_price") or ""),
        "date_on_sale_from": str(product.get("date_on_sale_from") or ""),
        "date_on_sale_to": str(product.get("date_on_sale_to") or ""),
        "description": _plain_text(str(product.get("description") or "")),
        "short_description": _plain_text(str(product.get("short_description") or "")),
        "manage_stock": "true" if product.get("manage_stock") else "false",
        "stock_quantity": "" if product.get("stock_quantity") is None else str(product.get("stock_quantity")),
        "stock_status": str(product.get("stock_status") or ""),
        "backorders": str(product.get("backorders") or ""),
        "sold_individually": "true" if product.get("sold_individually") else "false",
        "low_stock_amount": "" if product.get("low_stock_amount") is None else str(product.get("low_stock_amount")),
        "weight": str(product.get("weight") or ""),
        "length": str(dimensions.get("length") or ""),
        "width": str(dimensions.get("width") or ""),
        "height": str(dimensions.get("height") or ""),
        "shipping_class": str(product.get("shipping_class") or ""),
        "virtual": "true" if product.get("virtual") else "false",
        "downloadable": "true" if product.get("downloadable") else "false",
        "image_principal_id": str((product.get("images") or [{}])[0].get("id") or "") if product.get("images") else "",
        "image_galeria_ids": _join_image_ids((product.get("images") or [])[1:]),
        "image_nombres": _join_image_names(product.get("images") or []),
        "attr1_nombre": str((first_attribute or {}).get("name") or ""),
        "attr1_valores": _join_attribute_values(first_attribute),
        "attr2_nombre": str((second_attribute or {}).get("name") or ""),
        "attr2_valores": _join_attribute_values(second_attribute),
        "category_ids": _join_product_ids(product.get("categories")),
        "tag_ids": _join_product_ids(product.get("tags")),
        "upsell_ids": "; ".join(str(item) for item in (product.get("upsell_ids") or [])),
        "cross_sell_ids": "; ".join(str(item) for item in (product.get("cross_sell_ids") or [])),
        "meta_key": str(((product.get("meta_data") or [{}])[0]).get("key") or "") if product.get("meta_data") else "",
        "meta_value": str(((product.get("meta_data") or [{}])[0]).get("value") or "") if product.get("meta_data") else "",
        "images": "; ".join(part for part in (_join_image_ids(product.get("images")), _join_image_names(product.get("images"))) if part),
    }
    return mapping.get(field, "")


def _row_has_field(row: dict[str, str], field: str) -> bool:
    if field in BLANKABLE_FIELDS:
        return True
    if field == "images":
        return any(str(row.get(header, "")).strip() for header in FIELD_GROUPS["images"])
    return str(row.get(field, "")).strip() != ""


def _split_sku_values(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;,\n]+", str(value)) if item.strip()]


def _product_matcher(products: list[dict[str, Any]]):
    products_by_id = {int(product["id"]): product for product in products if str(product.get("id", "")).isdigit()}
    products_by_sku: dict[str, list[dict]] = {}
    for p in products:
        s = str(p.get("sku", "")).strip()
        if s:
            products_by_sku.setdefault(s, []).append(p)

    # Returns (product | None, identifier_str, sku_candidates_if_ambiguous)
    def matcher(row: dict[str, str]) -> tuple[dict | None, str, list[dict]]:
        sku = row.get("sku", "").strip()
        sku_tokens = _split_sku_values(sku)
        product_id = row.get("product_id", "").strip()
        if sku_tokens:
            candidate_groups = [products_by_sku.get(token, []) for token in sku_tokens]
            if any(not candidates for candidates in candidate_groups):
                return None, sku, []

            matching_ids = {
                int(candidate["id"])
                for candidate in candidate_groups[0]
                if str(candidate.get("id", "")).isdigit()
            }
            for candidates in candidate_groups[1:]:
                matching_ids &= {
                    int(candidate["id"])
                    for candidate in candidates
                    if str(candidate.get("id", "")).isdigit()
                }

            candidates = [
                candidate
                for candidate in candidate_groups[0]
                if str(candidate.get("id", "")).isdigit() and int(candidate["id"]) in matching_ids
            ]
            if len(candidates) == 1:
                return candidates[0], sku, []
            if len(candidates) > 1:
                return None, sku, candidates  # ambiguous — caller must resolve
            return None, sku, []
        if product_id.isdigit():
            return products_by_id.get(int(product_id)), product_id, []
        return None, sku or product_id or "sin identificador", []

    return matcher


async def _build_row_payload(
    row: dict[str, str],
    selected_fields: list[str],
    resolve_media_names=None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    selected = set(_expand_selected_fields(selected_fields))

    scalar_fields = selected - {
        "length",
        "width",
        "height",
        "image_principal_id",
        "image_galeria_ids",
        "image_nombres",
        "attr1_nombre",
        "attr1_valores",
        "attr2_nombre",
        "attr2_valores",
        "category_ids",
        "tag_ids",
        "upsell_ids",
        "cross_sell_ids",
        "meta_key",
        "meta_value",
    }

    for field in scalar_fields:
        if field not in row:
            continue
        raw_value = row.get(field, "")
        if raw_value == "" and field not in BLANKABLE_FIELDS:
            continue
        payload[field] = normalize_field_value(field, raw_value)

    if selected & {"length", "width", "height"}:
        dimensions = {
            key: str(row.get(key, "")).strip()
            for key in ("length", "width", "height")
            if key in selected
        }
        if dimensions:
            payload["dimensions"] = dimensions

    if selected & {"image_principal_id", "image_galeria_ids", "image_nombres"}:
        principal = str(row.get("image_principal_id", "")).strip()
        gallery = str(row.get("image_galeria_ids", "")).strip()
        image_names = str(row.get("image_nombres", "")).strip()
        image_ids: list[int] = []

        # Si el usuario llenó IDs explícitos, esos mandan sobre los nombres.
        if "image_principal_id" in selected and principal:
            image_ids.extend(_split_int_values(principal))
        if "image_galeria_ids" in selected and gallery:
            image_ids.extend(_split_int_values(gallery))

        # Solo resolvemos por nombre cuando no vienen IDs de principal o galería.
        if not image_ids and "image_nombres" in selected and image_names:
            if resolve_media_names is None:
                raise ValueError("No hay un resolvedor de media disponible para buscar imágenes por nombre")
            resolved_entries = await resolve_media_names(_split_values(image_names))
            unresolved = [entry["requested"] for entry in resolved_entries if not entry.get("matched") or not entry.get("item")]
            if unresolved:
                raise ValueError(f"No se encontraron estas imágenes: {', '.join(unresolved)}")
            image_ids.extend(int(entry["item"]["id"]) for entry in resolved_entries if entry.get("item"))

        if image_ids:
            deduped_image_ids: list[int] = []
            for image_id in image_ids:
                if image_id not in deduped_image_ids:
                    deduped_image_ids.append(image_id)
            payload["images"] = [{"id": image_id} for image_id in deduped_image_ids]
        elif not principal and not gallery and not image_names:
            # All image cells blank → clear product images
            payload["images"] = []

    if "category_ids" in selected:
        raw = str(row.get("category_ids", "")).strip()
        payload["categories"] = [{"id": item_id} for item_id in _split_int_values(raw)] if raw else []
    if "tag_ids" in selected:
        raw = str(row.get("tag_ids", "")).strip()
        payload["tags"] = [{"id": item_id} for item_id in _split_int_values(raw)] if raw else []

    if "upsell_ids" in selected:
        raw = str(row.get("upsell_ids", "")).strip()
        payload["upsell_ids"] = _split_int_values(raw) if raw else []

    if "cross_sell_ids" in selected:
        raw = str(row.get("cross_sell_ids", "")).strip()
        payload["cross_sell_ids"] = _split_int_values(raw) if raw else []

    if selected & {"attr1_nombre", "attr1_valores", "attr2_nombre", "attr2_valores"}:
        attributes: list[dict[str, Any]] = []
        for index in (1, 2):
            name_key = f"attr{index}_nombre"
            values_key = f"attr{index}_valores"
            name = str(row.get(name_key, "")).strip()
            values = _split_values(row.get(values_key, "")) if str(row.get(values_key, "")).strip() else []
            if name and values and ({name_key, values_key} & selected):
                attributes.append(
                    {
                        "name": name,
                        "position": index - 1,
                        "visible": True,
                        "variation": True,
                        "options": values,
                    }
                )
        if attributes:
            payload["attributes"] = attributes

    if selected & {"meta_key", "meta_value"}:
        meta_key = str(row.get("meta_key", "")).strip()
        if meta_key:
            payload["meta_data"] = [{"key": meta_key, "value": str(row.get("meta_value", "")).strip()}]

    return payload


async def generate_bulk_preview(
    *,
    rows: list[dict[str, str]],
    selected_fields: list[str],
    products: list[dict[str, Any]],
    product_id_overrides: dict[int, int] | None = None,
) -> dict[str, Any]:
    matcher = _product_matcher(products)
    products_by_id = {int(p["id"]): p for p in products if str(p.get("id", "")).isdigit()}
    preview_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows):
        # Apply user-chosen override for ambiguous rows
        if product_id_overrides and row_index in product_id_overrides:
            row = {**row, "product_id": str(product_id_overrides[row_index]), "sku": ""}
        product, identifier, sku_candidates = matcher(row)
        # If override given but not found via sku="" path, look up by id directly
        if product is None and product_id_overrides and row_index in product_id_overrides:
            product = products_by_id.get(product_id_overrides[row_index])
        changed_fields = [field for field in selected_fields if _row_has_field(row, field)]
        changes = [
            {
                "field": field,
                "before": "" if product is None else _current_field_value(product, field),
                "after": "; ".join(
                    row.get(header, "").strip()
                    for header in FIELD_GROUPS.get(field, [field])
                    if str(row.get(header, "")).strip()
                ),
            }
            for field in changed_fields
        ]
        preview_rows.append(
            {
                "row_index": row_index,
                "identifier": identifier,
                "product_found": product is not None,
                "product_id": None if not product else product.get("id"),
                "product_name": None if not product else product.get("name"),
                "changed_fields": changed_fields,
                "changes": changes,
                "row_data": row,
                "ambiguous": len(sku_candidates) > 0,
                "sku_candidates": [
                    {"id": c.get("id"), "name": c.get("name"), "sku": c.get("sku")}
                    for c in sku_candidates
                ],
            }
        )

    return {
        "preview_rows": preview_rows,
        "sample_rows": rows[:5],
        "total_rows": len(rows),
    }


async def build_bulk_update_summary(
    *,
    rows: list[dict[str, str]],
    selected_fields: list[str],
    products: list[dict[str, Any]],
    woo_service,
    resolve_media_names=None,
    selected_row_indexes: list[int] | None = None,
) -> tuple[UpdateSummary, list[dict[str, Any]]]:
    update_payloads, errors, selected_rows = await prepare_bulk_update_payloads(
        rows=rows,
        selected_fields=selected_fields,
        products=products,
        resolve_media_names=resolve_media_names,
        selected_row_indexes=selected_row_indexes,
    )

    updated = 0
    updated_products: list[dict[str, Any]] = []
    for start in range(0, len(update_payloads), 100):
        batch = update_payloads[start : start + 100]
        batch_payload = [{key: value for key, value in item.items() if not key.startswith("_")} for item in batch]
        try:
            result = await woo_service.batch_update(batch_payload)
            # WC returns ALL sent products in "update" — including failures as error objects.
            # A failed product has a "code" field (WC error code) instead of product fields.
            # We must separate good responses from error responses.
            for item in result.get("update", []):
                if "code" in item and "message" in item:
                    # WC-level error for this specific product
                    identifier = str(item.get("data", {}).get("id") or item.get("id") or "desconocido")
                    errors.append({"identifier": identifier, "error": item.get("message", "Error de WooCommerce")})
                elif item.get("id"):
                    updated += 1
                    updated_products.append(item)
                else:
                    # Unexpected shape — count as error, don't corrupt backup
                    errors.append({"identifier": "desconocido", "error": f"Respuesta inesperada de WooCommerce: {str(item)[:120]}"})
        except Exception as exc:
            for item in batch:
                errors.append({"identifier": str(item.get("_identifier") or item.get("id", "?")), "error": str(exc)})

    return UpdateSummary(updated=updated, failed=len(errors), errors=errors, total_rows=len(selected_rows)), updated_products


async def prepare_bulk_update_payloads(
    *,
    rows: list[dict[str, str]],
    selected_fields: list[str],
    products: list[dict[str, Any]],
    resolve_media_names=None,
    selected_row_indexes: list[int] | None = None,
    on_row_progress=None,
    product_id_overrides: dict[int, int] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, str]], list[tuple[int, dict[str, str]]]]:
    matcher = _product_matcher(products)
    update_payloads: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    selected_indexes = set(range(len(rows))) if selected_row_indexes is None else set(selected_row_indexes)

    if not selected_indexes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes seleccionar al menos una fila para aplicar cambios")

    selected_rows = [(row_index, row) for row_index, row in enumerate(rows) if row_index in selected_indexes]

    for row_index, row in selected_rows:
        # Apply user-chosen product when SKU was ambiguous
        if product_id_overrides and row_index in product_id_overrides:
            row = {**row, "product_id": str(product_id_overrides[row_index]), "sku": ""}
        product, identifier, sku_candidates = matcher(row)
        if on_row_progress is not None:
            await on_row_progress(
                stage="preparing",
                row_index=row_index,
                identifier=identifier,
                product=None if not product else product.get("name"),
            )
        if not product:
            if sku_candidates:
                errors.append({"identifier": identifier, "error": f"SKU ambiguo: {len(sku_candidates)} productos comparten este SKU. Selecciona uno manualmente."})
            else:
                errors.append({"identifier": identifier, "error": "Producto no encontrado"})
            continue
        try:
            payload = await _build_row_payload(row, selected_fields, resolve_media_names=resolve_media_names)
            payload.pop("sku", None)
            if not payload:
                continue
            payload["id"] = product["id"]
            payload["_identifier"] = identifier
            payload["_row_index"] = row_index
            payload["_product_name"] = product.get("name")
            update_payloads.append(payload)
            if on_row_progress is not None:
                await on_row_progress(
                    stage="prepared",
                    row_index=row_index,
                    identifier=identifier,
                    product=product.get("name"),
                )
        except Exception as exc:
            errors.append({"identifier": identifier, "error": f"Invalid value: {exc}"})
            if on_row_progress is not None:
                await on_row_progress(
                    stage="failed",
                    row_index=row_index,
                    identifier=identifier,
                    product=product.get("name"),
                    error=str(exc),
                )

    return update_payloads, errors, selected_rows


async def generate_preview(
    *,
    rows: list[dict[str, str]],
    id_column: str,
    value_column: str,
    id_type: str,
    wc_field: str,
    products: list[dict[str, Any]],
) -> dict[str, Any]:
    identifiers = [row.get(id_column, "").strip() for row in rows if row.get(id_column)]

    if id_type == "product_id":
        products_by_id = {int(product["id"]): product for product in products}

        def matcher(identifier: str):
            return products_by_id.get(int(identifier)) if identifier.isdigit() else None
    else:
        products_by_sku = {str(product.get("sku", "")): product for product in products if product.get("sku")}

        def matcher(identifier: str):
            return products_by_sku.get(identifier)

    preview_rows: list[dict[str, Any]] = []
    for row in rows:
        identifier = row.get(id_column, "").strip()
        if not identifier:
            continue
        product = matcher(identifier)
        preview_row = PreviewRow(
            identifier=identifier,
            current_value=None if not product else product.get(wc_field),
            new_value=row.get(value_column, ""),
            product_found=product is not None,
            product_id=None if not product else product.get("id"),
            product_name=None if not product else product.get("name"),
        )
        preview_rows.append(preview_row.model_dump())

    return {
        "preview_rows": preview_rows,
        "sample_rows": rows[:5],
        "total_rows": len(rows),
    }


async def build_update_summary(
    *,
    rows: list[dict[str, str]],
    id_column: str,
    value_column: str,
    id_type: str,
    wc_field: str,
    products: list[dict[str, Any]],
    woo_service,
) -> tuple[UpdateSummary, list[dict[str, Any]]]:
    identifiers = [row.get(id_column, "").strip() for row in rows if row.get(id_column)]

    if id_type == "product_id":
        products_by_id = {int(product["id"]): product for product in products}

        def matcher(identifier: str):
            return products_by_id.get(int(identifier)) if identifier.isdigit() else None
    else:
        products_by_sku = {str(product.get("sku", "")): product for product in products if product.get("sku")}

        def matcher(identifier: str):
            return products_by_sku.get(identifier)

    update_payloads: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []

    for row in rows:
        identifier = row.get(id_column, "").strip()
        if not identifier:
            continue
        product = matcher(identifier)
        if not product:
            errors.append({"identifier": identifier, "error": "Product not found"})
            continue
        try:
            update_payloads.append({"id": product["id"], wc_field: normalize_field_value(wc_field, row.get(value_column, ""))})
        except Exception as exc:
            errors.append({"identifier": identifier, "error": f"Invalid value: {exc}"})

    updated = 0
    updated_products: list[dict[str, Any]] = []
    for start in range(0, len(update_payloads), 100):
        batch = update_payloads[start : start + 100]
        try:
            result = await woo_service.batch_update(batch)
            batch_updates = result.get("update", [])
            updated += len(batch_updates)
            updated_products.extend(batch_updates)
            for failed in result.get("not_updated", []):
                errors.append(
                    {
                        "identifier": str(failed.get("id", "unknown")),
                        "error": failed.get("error", {}).get("message", "Update failed"),
                    }
                )
        except Exception as exc:
            for item in batch:
                errors.append({"identifier": str(item["id"]), "error": str(exc)})

    return UpdateSummary(updated=updated, failed=len(errors), errors=errors, total_rows=len(rows)), updated_products
