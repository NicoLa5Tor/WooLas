from __future__ import annotations

import asyncio
import html
import uuid
from io import BytesIO
from pathlib import Path, PurePosixPath
import re
from uuid import UUID
from difflib import SequenceMatcher

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from core.config import settings
from repositories import import_draft as import_draft_repository
from schemas.imports import (
    ImportApplyResult,
    ImportExpectedField,
    ImportFormatAnalysis,
    ImportHeaderSuggestion,
    ImportImagesJobRowStatus,
    ImportImagesJobStatus,
)
from services import excel as excel_service
from services import media as media_service

DARK = "1E293B"
MID = "334155"
BLUE = "1D4ED8"
GREEN = "065F46"
PURPLE = "6B21A8"
TEAL = "0F766E"
ORANGE = "C2410C"
ALT = "F8FAFC"
WHITE = "FFFFFF"

thin = Side(style="thin", color="CBD5E1")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMN_DEFINITIONS = [
    {"key": "product_id", "label": "ID del producto", "group": "IDENTIFICACION", "color": DARK, "required": True},
    {"key": "sku", "label": "SKU", "group": "IDENTIFICACION", "color": DARK, "required": True},
    {"key": "name", "label": "Nombre", "group": "GENERAL", "color": MID, "required": True},
    {"key": "slug", "label": "Slug", "group": "GENERAL", "color": MID, "required": False},
    {"key": "type", "label": "Tipo", "group": "GENERAL", "color": MID, "required": True},
    {"key": "status", "label": "Estado", "group": "GENERAL", "color": MID, "required": True},
    {"key": "featured", "label": "Destacado", "group": "GENERAL", "color": MID, "required": False},
    {"key": "catalog_visibility", "label": "Visibilidad catalogo", "group": "GENERAL", "color": MID, "required": False},
    {"key": "regular_price", "label": "Precio regular", "group": "PRECIOS", "color": GREEN, "required": True},
    {"key": "sale_price", "label": "Precio oferta", "group": "PRECIOS", "color": GREEN, "required": False},
    {"key": "date_on_sale_from", "label": "Oferta desde", "group": "PRECIOS", "color": GREEN, "required": False},
    {"key": "date_on_sale_to", "label": "Oferta hasta", "group": "PRECIOS", "color": GREEN, "required": False},
    {"key": "description", "label": "Descripcion completa", "group": "DESCRIPCIONES", "color": TEAL, "required": False},
    {"key": "short_description", "label": "Descripcion corta", "group": "DESCRIPCIONES", "color": TEAL, "required": False},
    {"key": "manage_stock", "label": "Gestionar stock", "group": "INVENTARIO", "color": BLUE, "required": False},
    {"key": "stock_quantity", "label": "Cantidad stock", "group": "INVENTARIO", "color": BLUE, "required": False},
    {"key": "stock_status", "label": "Estado stock", "group": "INVENTARIO", "color": BLUE, "required": False},
    {"key": "backorders", "label": "Backorders", "group": "INVENTARIO", "color": BLUE, "required": False},
    {"key": "sold_individually", "label": "Venta individual", "group": "INVENTARIO", "color": BLUE, "required": False},
    {"key": "low_stock_amount", "label": "Alerta stock bajo", "group": "INVENTARIO", "color": BLUE, "required": False},
    {"key": "weight", "label": "Peso", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "length", "label": "Largo", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "width", "label": "Ancho", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "height", "label": "Alto", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "shipping_class", "label": "Clase de envio", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "virtual", "label": "Virtual", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "downloadable", "label": "Descargable", "group": "ENVIO", "color": PURPLE, "required": False},
    {"key": "image_principal_id", "label": "ID imagen principal", "group": "IMAGENES", "color": ORANGE, "required": False},
    {"key": "image_galeria_ids", "label": "IDs galeria", "group": "IMAGENES", "color": ORANGE, "required": False},
    {"key": "image_nombres", "label": "Nombres imagenes", "group": "IMAGENES", "color": ORANGE, "required": False},
    {"key": "attr1_nombre", "label": "Atributo 1 nombre", "group": "ATRIBUTOS", "color": "B45309", "required": False},
    {"key": "attr1_valores", "label": "Atributo 1 valores", "group": "ATRIBUTOS", "color": "B45309", "required": False},
    {"key": "attr2_nombre", "label": "Atributo 2 nombre", "group": "ATRIBUTOS", "color": "B45309", "required": False},
    {"key": "attr2_valores", "label": "Atributo 2 valores", "group": "ATRIBUTOS", "color": "B45309", "required": False},
    {"key": "category_ids", "label": "IDs categorias", "group": "RELACIONADOS", "color": "0369A1", "required": False},
    {"key": "tag_ids", "label": "IDs etiquetas", "group": "RELACIONADOS", "color": "0369A1", "required": False},
    {"key": "upsell_ids", "label": "IDs upsells", "group": "RELACIONADOS", "color": "0369A1", "required": False},
    {"key": "cross_sell_ids", "label": "IDs cross-sells", "group": "RELACIONADOS", "color": "0369A1", "required": False},
    {"key": "meta_key", "label": "Meta key", "group": "META DATA", "color": "4B5563", "required": False},
    {"key": "meta_value", "label": "Meta value", "group": "META DATA", "color": "4B5563", "required": False},
]

HEADER_LABELS = {item["key"]: item["label"] for item in COLUMN_DEFINITIONS}
HTML_TAG_RE = re.compile(r"<[^>]+>")
GROUP_RANGES = [
    (1, 2, "IDENTIFICACION", DARK),
    (3, 8, "GENERAL", MID),
    (9, 12, "PRECIOS", GREEN),
    (13, 14, "DESCRIPCIONES", TEAL),
    (15, 20, "INVENTARIO", BLUE),
    (21, 27, "ENVIO", PURPLE),
    (28, 30, "IMAGENES", ORANGE),
    (31, 34, "ATRIBUTOS", "B45309"),
    (35, 38, "RELACIONADOS", "0369A1"),
    (39, 40, "META DATA", "4B5563"),
]

import_apply_jobs: dict[str, ImportImagesJobStatus] = {}


def _draft_directory(tenant_id: UUID) -> Path:
    return settings.imports_dir / str(tenant_id)


def _draft_file_path(tenant_id: UUID, draft_id: UUID, original_filename: str) -> Path:
    safe_name = Path(original_filename or "import.xlsx").name
    return _draft_directory(tenant_id) / f"{draft_id}_{safe_name}"


def _serialize_draft(record) -> dict:
    return {
        "id": str(record.id),
        "tenant_id": str(record.tenant_id),
        "original_filename": record.original_filename,
        "headers": record.headers,
        "sample_rows": record.sample_rows,
        "row_count": record.row_count,
        "created_at": record.created_at.isoformat(),
        "updated_at": record.updated_at.isoformat(),
    }


def _parse_xlsx_payload(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]], int]:
    headers, rows = excel_service.parse_excel(file_bytes)
    return headers, rows[:5], len(rows)


def _parse_xlsx_source_payload(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]], int]:
    headers, rows = excel_service.parse_excel(file_bytes, normalize_headers=False)
    return headers, rows[:5], len(rows)


def _write_file(path: Path, file_bytes: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(file_bytes)


def _remove_file(path_str: str) -> None:
    path = Path(path_str)
    try:
        if path.exists():
            path.unlink()
    except OSError:
        pass


def _expected_fields() -> list[ImportExpectedField]:
    return [ImportExpectedField(key=item["key"], label=item["label"], required=bool(item["required"])) for item in COLUMN_DEFINITIONS]


def _clean_header_label(value: str) -> str:
    cleaned = value.strip()
    if cleaned.startswith("* "):
        cleaned = cleaned[2:].strip()
    return cleaned


def _analyze_headers(source_headers: list[str], normalized_headers: list[str]) -> ImportFormatAnalysis:
    expected_labels = [item["label"] for item in COLUMN_DEFINITIONS]
    required_labels = [item["label"] for item in COLUMN_DEFINITIONS if item["required"]]
    typo_suggestions: list[ImportHeaderSuggestion] = []
    unknown_headers: list[str] = []
    cleaned_source_headers = [_clean_header_label(header) for header in source_headers]

    for original_header, header in zip(source_headers, cleaned_source_headers, strict=False):
        if header in expected_labels:
            continue
        scores = sorted(
            ((label, SequenceMatcher(None, header.lower(), label.lower()).ratio()) for label in expected_labels),
            key=lambda item: item[1],
            reverse=True,
        )
        best_label, best_score = scores[0] if scores else ("", 0.0)
        if best_score >= 0.72:
            typo_suggestions.append(ImportHeaderSuggestion(provided=original_header, expected=best_label, confidence=round(best_score, 2)))
        else:
            unknown_headers.append(original_header)

    present_labels = set(cleaned_source_headers)
    suggested_labels = {item.expected for item in typo_suggestions}
    missing_required = [label for label in required_labels if label not in present_labels and label not in suggested_labels]

    if not typo_suggestions and not unknown_headers and not missing_required:
        status = "ready"
    elif typo_suggestions and not unknown_headers and not missing_required:
        status = "typos"
    else:
        status = "refactor_required"

    return ImportFormatAnalysis(
        status=status,
        source_headers=source_headers,
        normalized_headers=normalized_headers,
        typo_suggestions=typo_suggestions,
        missing_required=missing_required,
        unknown_headers=unknown_headers,
        expected_fields=_expected_fields(),
    )


def _build_draft_response(record) -> dict:
    file_bytes = Path(record.storage_path).read_bytes() if Path(record.storage_path).exists() else b""
    source_headers, _, _ = _parse_xlsx_source_payload(file_bytes) if file_bytes else ([], [], 0)
    analysis = _analyze_headers(source_headers, record.headers or [])
    payload = _serialize_draft(record)
    payload["source_headers"] = source_headers
    payload["format_analysis"] = analysis.model_dump()
    payload["sample_rows"] = record.sample_rows or []
    return payload


def _format_bool(value: object) -> str:
    return "true" if bool(value) else "false"


def _clean_text(value: object) -> str:
    """Decode HTML entities (&#8211; → –, &amp; → &, etc.) and strip."""
    if value is None or value == "":
        return ""
    return html.unescape(str(value)).strip()


def _join_ids(items: list[object]) -> str:
    return ";".join(str(item) for item in items if item not in {None, ""})


def _join_taxonomy_ids(items: list[dict]) -> str:
    return ";".join(str(item.get("id")) for item in items if item.get("id") is not None)


def _join_attribute_options(attribute: dict | None) -> str:
    if not attribute:
        return ""
    return ";".join(_clean_text(option) for option in attribute.get("options", []) if str(option).strip())


def _image_name(image: dict) -> str:
    raw = str(image.get("name") or "").strip()
    if raw:
        return html.unescape(raw)
    return (
        PurePosixPath(str(image.get("src") or "")).name
        or f"imagen-{image.get('id', '')}"
    )


def _html_to_plain_text(value: object) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = re.sub(r"</p>\s*<p>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = HTML_TAG_RE.sub("", text)
    text = html.unescape(text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def _autosize_columns(sheet, *, min_width: int = 12, max_width: int = 48) -> None:
    for column_cells in sheet.columns:
        values = [str(cell.value).strip() for cell in column_cells if cell.value is not None]
        if not values:
            continue
        real_cell = next((cell for cell in column_cells if hasattr(cell, "column_letter")), None)
        if real_cell is None:
            continue
        width = max(len(value) for value in values) + 2
        sheet.column_dimensions[real_cell.column_letter].width = max(min_width, min(width, max_width))


def _product_to_template_row(product: dict) -> list[str]:
    dimensions = product.get("dimensions") or {}
    images = product.get("images") or []
    attributes = product.get("attributes") or []
    attr1 = attributes[0] if len(attributes) > 0 else None
    attr2 = attributes[1] if len(attributes) > 1 else None
    meta = next((item for item in product.get("meta_data", []) if str(item.get("key") or "").strip()), None)

    return [
        str(product.get("id") or ""),
        _clean_text(product.get("sku")),
        _clean_text(product.get("name")),
        _clean_text(product.get("slug")),
        _clean_text(product.get("type")),
        _clean_text(product.get("status")),
        _format_bool(product.get("featured")),
        _clean_text(product.get("catalog_visibility")),
        str(product.get("regular_price") or ""),
        str(product.get("sale_price") or ""),
        str(product.get("date_on_sale_from") or ""),
        str(product.get("date_on_sale_to") or ""),
        _html_to_plain_text(product.get("description")),
        _html_to_plain_text(product.get("short_description")),
        _format_bool(product.get("manage_stock")),
        "" if product.get("stock_quantity") is None else str(product.get("stock_quantity")),
        _clean_text(product.get("stock_status")),
        _clean_text(product.get("backorders")),
        _format_bool(product.get("sold_individually")),
        "" if product.get("low_stock_amount") is None else str(product.get("low_stock_amount")),
        str(product.get("weight") or ""),
        str(dimensions.get("length") or ""),
        str(dimensions.get("width") or ""),
        str(dimensions.get("height") or ""),
        _clean_text(product.get("shipping_class")),
        _format_bool(product.get("virtual")),
        _format_bool(product.get("downloadable")),
        str(images[0].get("id") or "") if images else "",
        _join_ids([image.get("id") for image in images[1:]]),
        ";".join(_image_name(image) for image in images if image),
        _clean_text(attr1.get("name")) if attr1 else "",
        _join_attribute_options(attr1),
        _clean_text(attr2.get("name")) if attr2 else "",
        _join_attribute_options(attr2),
        _join_taxonomy_ids(product.get("categories", [])),
        _join_taxonomy_ids(product.get("tags", [])),
        _join_ids(product.get("upsell_ids", [])),
        _join_ids(product.get("cross_sell_ids", [])),
        str(meta.get("key") or "") if meta else "",
        str(meta.get("value") or "") if meta else "",
    ]


def _build_template_workbook(rows: list[list[str]], title: str, *, highlight_first_row: bool, categories: list[dict] | None = None, media_records=None) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "productos"
    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:AN1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=13, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for start_col, end_col, label, color in GROUP_RANGES:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        c = ws.cell(row=2, column=start_col)
        c.value = label
        c.font = Font(bold=True, color=WHITE, size=8, name="Arial")
        c.fill = PatternFill("solid", start_color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start_col, end_col + 1):
            ws.cell(row=2, column=col).fill = PatternFill("solid", start_color=color)
    ws.row_dimensions[2].height = 18

    for col, item in enumerate(COLUMN_DEFINITIONS, 1):
        c = ws.cell(row=3, column=col)
        c.value = ("* " if item["required"] else "") + item["label"]
        c.font = Font(bold=True, color=WHITE, size=8, name="Arial")
        c.fill = PatternFill("solid", start_color=item["color"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[3].height = 30

    for row_index, row in enumerate(rows, start=4):
        is_highlighted = highlight_first_row and row_index == 4
        bg = "FEF9C3" if is_highlighted else (WHITE if row_index % 2 == 0 else ALT)
        font_color = "92400E" if is_highlighted else DARK
        italic = is_highlighted
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_index, column=col, value=val)
            c.font = Font(size=9, name="Arial", color=font_color, italic=italic)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = brd

    wi = workbook.create_sheet("instrucciones")
    wi.sheet_view.showGridLines = False

    wi.merge_cells("A1:D1")
    c = wi["A1"]
    c.value = "INSTRUCCIONES DE USO"
    c.font = Font(bold=True, size=13, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    wi.row_dimensions[1].height = 28

    for col, header in enumerate(["Campo", "Obligatorio", "Valores aceptados", "Notas"], 1):
        c = wi.cell(row=2, column=col, value=header)
        c.font = Font(bold=True, color=WHITE, size=9, name="Arial")
        c.fill = PatternFill("solid", start_color=MID)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd

    instrucciones = [
        ("ID del producto", "Si (o SKU)", "Numero entero", "ID de WooCommerce. Usar para empatar producto existente."),
        ("SKU", "Si (o ID)", "Texto libre", "Codigo unico del producto. Preferido sobre el ID."),
        ("Nombre", "Si", "Texto libre", "Nombre visible del producto en la tienda."),
        ("Slug", "No", "Texto sin espacios", "Se genera automaticamente si se deja vacio."),
        ("Tipo", "Si", "simple | variable | grouped | external", ""),
        ("Estado", "Si", "publish | draft | pending | private", ""),
        ("Destacado", "No", "true | false", ""),
        ("Visibilidad catalogo", "No", "visible | catalog | search | hidden", ""),
        ("Precio regular", "Si", "Numero", "Ej: 129900"),
        ("Precio oferta", "No", "Numero", ""),
        ("Oferta desde / hasta", "No", "ISO 8601", "Ej: 2026-04-25T08:00:00Z"),
        ("Gestionar stock", "No", "true | false", ""),
        ("Cantidad stock", "No", "Numero entero", "Solo aplica si gestionas stock."),
        ("Estado stock", "No", "instock | outofstock | onbackorder", ""),
        ("Backorders", "No", "no | notify | yes", ""),
        ("Peso / Largo / Ancho / Alto", "No", "Numero decimal", "Dimensiones del producto."),
        ("IDs imagenes", "No", "IDs separados por ;", "La principal va aparte de la galeria."),
        ("Atributos", "No", "Nombre + valores separados por ;", "Actualmente la plantilla exporta 2 atributos visibles."),
        ("IDs categorias / etiquetas", "No", "IDs separados por ;", ""),
        ("IDs upsells / cross-sells", "No", "IDs separados por ;", ""),
        ("Meta key / value", "No", "Texto libre", "Se exporta la primera meta_data util."),
    ]
    for i, row in enumerate(instrucciones):
        bg = WHITE if i % 2 == 0 else ALT
        for col, val in enumerate(row, 1):
            c = wi.cell(row=i + 3, column=col, value=val)
            c.font = Font(size=9, name="Arial", color=DARK)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = brd
        wi.row_dimensions[i + 3].height = 28

    wr = workbook.create_sheet("referencia")
    wr.sheet_view.showGridLines = False
    wr.merge_cells("A1:H1")
    c = wr["A1"]
    c.value = "REFERENCIA DE VALORES VALIDOS"
    c.font = Font(bold=True, size=13, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    wr.row_dimensions[1].height = 28

    refs = [
        ("Tipo", ["simple", "variable", "grouped", "external"]),
        ("Estado", ["publish", "draft", "pending", "private"]),
        ("Visibilidad catalogo", ["visible", "catalog", "search", "hidden"]),
        ("Estado stock", ["instock", "outofstock", "onbackorder"]),
        ("Backorders", ["no", "notify", "yes"]),
        ("Booleanos", ["true", "false"]),
        ("Separador listas", ["; (punto y coma)"]),
        ("Formato fechas", ["ISO 8601: 2026-04-25T08:00:00Z"]),
    ]
    for col_ref, (field, values) in enumerate(refs, 1):
        c = wr.cell(row=2, column=col_ref, value=field)
        c.font = Font(bold=True, color=WHITE, size=9, name="Arial")
        c.fill = PatternFill("solid", start_color=MID)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
        wr.column_dimensions[get_column_letter(col_ref)].width = 28
        for idx, val in enumerate(values):
            c = wr.cell(row=3 + idx, column=col_ref, value=val)
            c.font = Font(size=9, name="Arial", color=DARK)
            c.fill = PatternFill("solid", start_color=WHITE if idx % 2 == 0 else ALT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd

    _autosize_columns(ws, min_width=12, max_width=44)
    _autosize_columns(wi, min_width=16, max_width=56)
    _autosize_columns(wr, min_width=18, max_width=32)

    # --- Hoja categorias (datos en vivo desde WC) ---
    cat_rows = [
        [
            str(cat.get("id") or ""),
            str(cat.get("name") or ""),
            str(cat.get("slug") or ""),
            str(cat.get("parent") or "0"),
            str(cat.get("description") or ""),
            str(cat.get("count") or "0"),
        ]
        for cat in (categories or [])
    ]
    _add_catalog_sheet(
        workbook,
        sheet_name="categorias",
        title="CATEGORIAS — usa el ID en la columna 'IDs categorias' del Excel de productos",
        headers=[
            ("ID", "0369A1"),
            ("Nombre", "0369A1"),
            ("Slug", "0369A1"),
            ("ID categoria padre (0 = raiz)", "0369A1"),
            ("Descripcion", "0369A1"),
            ("Productos", "0369A1"),
        ],
        rows=cat_rows,
    )

    # --- Hoja imagenes (desde media index local) ---
    img_rows = [
        [
            str(rec.media_id),
            str(rec.filename or ""),
            str(rec.title or ""),
            str(rec.url or ""),
            str(rec.thumbnail or ""),
        ]
        for rec in (media_records or [])
    ]
    _add_catalog_sheet(
        workbook,
        sheet_name="imagenes",
        title="IMAGENES (media index) — usa el ID o el nombre en las columnas de imagenes",
        headers=[
            ("ID WordPress", ORANGE),
            ("Nombre archivo", ORANGE),
            ("Titulo", ORANGE),
            ("URL completa", ORANGE),
            ("Thumbnail", ORANGE),
        ],
        rows=img_rows,
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_import_draft(
    db: Session,
    *,
    tenant_id: UUID,
    user_id: UUID | None,
    filename: str,
    file_bytes: bytes,
) -> dict:
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes seleccionar un archivo .xlsx")

    try:
        headers, sample_rows, row_count = _parse_xlsx_payload(file_bytes)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"No se pudo leer el Excel: {exc}")

    existing = import_draft_repository.get_import_draft_for_tenant(db, tenant_id)
    if existing is not None:
        _remove_file(existing.storage_path)

    draft_id = uuid.uuid4()
    storage_path = _draft_file_path(tenant_id, draft_id, filename)
    _write_file(storage_path, file_bytes)

    record = import_draft_repository.upsert_import_draft(
        db,
        tenant_id=tenant_id,
        created_by_user_id=user_id,
        original_filename=filename or "import.xlsx",
        storage_path=str(storage_path),
        headers=headers,
        sample_rows=sample_rows,
        row_count=row_count,
    )
    db.commit()
    return _build_draft_response(record)


def get_current_import_draft(db: Session, tenant_id: UUID) -> dict | None:
    record = import_draft_repository.get_import_draft_for_tenant(db, tenant_id)
    return None if record is None else _build_draft_response(record)


def require_current_import_draft(db: Session, tenant_id: UUID):
    record = import_draft_repository.get_import_draft_for_tenant(db, tenant_id)
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No hay un Excel guardado para este cliente")
    return record


def get_import_draft(db: Session, tenant_id: UUID, draft_id: UUID):
    record = import_draft_repository.get_import_draft_by_id(db, tenant_id, draft_id)
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import draft not found")
    return record


def get_import_draft_bytes(db: Session, tenant_id: UUID, draft_id: UUID) -> bytes:
    record = get_import_draft(db, tenant_id, draft_id)
    path = Path(record.storage_path)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Imported file not found")
    return path.read_bytes()


def get_current_import_draft_bytes(db: Session, tenant_id: UUID) -> tuple[dict, bytes]:
    record = require_current_import_draft(db, tenant_id)
    path = Path(record.storage_path)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Imported file not found")
    return _build_draft_response(record), path.read_bytes()



def _add_catalog_sheet(
    workbook,
    sheet_name: str,
    title: str,
    headers: list[tuple[str, str]],  # (label, color)
    rows: list[list[str]],
) -> None:
    """Generic helper to add a reference sheet with title row, header row, and data rows."""
    ws = workbook.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n_cols = len(headers)
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=12, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for col, (label, color) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = Font(bold=True, color=WHITE, size=9, name="Arial")
        c.fill = PatternFill("solid", start_color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    ws.row_dimensions[2].height = 22

    for i, row in enumerate(rows):
        bg = WHITE if i % 2 == 0 else ALT
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i + 3, column=col, value=val)
            c.font = Font(size=9, name="Arial", color=DARK)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = brd
        ws.row_dimensions[i + 3].height = 20

    if not rows:
        ws.cell(row=3, column=1, value="(Sin datos en el backup actual)")

    _autosize_columns(ws, min_width=10, max_width=52)


def build_import_template(
    categories: list[dict] | None = None,
    media_records=None,  # list[MediaIndexRecord]
) -> tuple[str, bytes]:
    sample = [
        "12345", "03330154", "Producto ejemplo", "producto-ejemplo",
        "variable", "publish", "true", "visible",
        "129900", "119900", "2026-04-25T08:00:00Z", "2026-04-30T23:59:59Z",
        "<p>Descripcion completa.</p>", "<p>Descripcion corta.</p>",
        "true", "10", "instock", "no", "false", "2",
        "1.2", "20", "10", "5", "hogar", "false", "false",
        "501", "502;503", "03330154-Nombre-Marca-1.jpg",
        "Color", "Rojo;Azul", "Talla", "S;M;L",
        "12;15", "3;9", "220;221", "330;331",
        "_origen_excel", "campana_mayo",
    ]
    workbook_bytes = _build_template_workbook(
        [sample], "PLANTILLA DE IMPORTACION - WOOCOMMERCE", highlight_first_row=True,
        categories=categories or [],
        media_records=media_records or [],
    )
    return "plantilla_productos_woocommerce.xlsx", workbook_bytes


def build_products_export(products: list[dict]) -> tuple[str, bytes]:
    rows = [_product_to_template_row(product) for product in products]
    return "productos_woocommerce.xlsx", _build_template_workbook(rows, "EXPORTACION DE PRODUCTOS - WOOCOMMERCE", highlight_first_row=False)


def build_media_export(media_records) -> tuple[str, bytes]:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "imagenes"
    ws.sheet_view.showGridLines = False

    title_text = "BIBLIOTECA DE IMÁGENES - WOOCOMMERCE"
    headers = [
        ("ID WordPress", ORANGE),
        ("Nombre archivo", ORANGE),
        ("Titulo", ORANGE),
        ("Slug", ORANGE),
        ("URL completa", ORANGE),
        ("Thumbnail", ORANGE),
        ("Subida", ORANGE),
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, (label, color) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = brd
    ws.row_dimensions[2].height = 24

    for row_offset, rec in enumerate(media_records or [], start=3):
        values = [
            str(rec.media_id),
            _clean_text(rec.filename),
            _clean_text(rec.title),
            _clean_text(rec.slug),
            str(rec.url or ""),
            str(rec.thumbnail or ""),
            rec.uploaded_at.isoformat() if rec.uploaded_at else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = brd
            if row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT)

    if not media_records:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        empty_cell = ws.cell(row=3, column=1, value="(Sin imágenes en el índice — sincroniza la biblioteca primero)")
        empty_cell.font = Font(italic=True, color="64748B")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"
    _autosize_columns(ws, min_width=12, max_width=60)

    output = BytesIO()
    workbook.save(output)
    return "imagenes_woocommerce.xlsx", output.getvalue()


def _canonical_row_to_template_row(row: dict[str, str]) -> list[str]:
    return [str(row.get(item["key"]) or "") for item in COLUMN_DEFINITIONS]


def refactor_import_draft(
    db: Session,
    *,
    tenant_id: UUID,
    user_id: UUID | None,
    mapping: dict[str, str | None],
) -> dict:
    draft, file_bytes = get_current_import_draft_bytes(db, tenant_id)
    source_headers, rows = excel_service.parse_excel(file_bytes, normalize_headers=False)
    if not source_headers:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no tiene columnas para refactorizar")

    canonical_rows: list[dict[str, str]] = []
    for row in rows:
        canonical_row: dict[str, str] = {}
        for item in COLUMN_DEFINITIONS:
            source_header = mapping.get(item["key"])
            canonical_row[item["key"]] = row.get(source_header or "", "") if source_header else ""
        canonical_rows.append(canonical_row)

    workbook_bytes = _build_template_workbook(
        [_canonical_row_to_template_row(row) for row in canonical_rows],
        "EXCEL REFACTORIZADO - WOOCOMMERCE",
        highlight_first_row=False,
    )
    return save_import_draft(
        db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"refactorizado_{draft['original_filename']}",
        file_bytes=workbook_bytes,
    )


async def preview_import_draft(
    *,
    db: Session,
    tenant_id: UUID,
    selected_fields: list[str],
    products: list[dict],
    product_id_overrides: dict[int, int] | None = None,
):
    _, file_bytes = get_current_import_draft_bytes(db, tenant_id)
    headers, rows = excel_service.parse_excel(file_bytes)
    excel_service.validate_selected_fields(headers, selected_fields)
    preview = await excel_service.generate_bulk_preview(
        rows=rows,
        selected_fields=selected_fields,
        products=products,
        product_id_overrides=product_id_overrides,
    )
    return {"headers": headers, **preview}


def _row_split_values(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;,\n]+", str(value)) if item.strip()]


def _row_split_int_values(value: str) -> list[int]:
    parsed: list[int] = []
    for item in _row_split_values(value):
        try:
            parsed.append(int(float(item)))
        except ValueError:
            pass
    return parsed


async def _run_import_images_job(
    *,
    job_id: str,
    tenant_id: UUID,
    backup_id: UUID,
    rows: list[dict[str, str]],
    selected_row_indexes: list[int] | None,
    products: list[dict],
    credentials: dict[str, str],
):
    from core.database import SessionLocal
    from repositories import backup as backup_repository
    from services import backup as backup_service
    from services.woocommerce import WooCommerceService

    job = import_apply_jobs[job_id]
    woo_service = WooCommerceService(**credentials)
    updated_products: list[dict] = []
    result = ImportApplyResult()
    job_db = None

    try:
        job.status = "running"
        job.stage = "preparing"
        job.progress = 2
        job_db = SessionLocal()

        # 1. Construir lista de productos a procesar
        selected_indexes = set(range(len(rows))) if selected_row_indexes is None else set(selected_row_indexes)
        matcher = excel_service._product_matcher(products)

        matched: list[tuple[ImportImagesJobRowStatus, dict[str, str], dict]] = []
        for row_index, row in enumerate(rows):
            if row_index not in selected_indexes:
                continue
            product, identifier, sku_candidates = matcher(row)
            if product is None:
                result.failed += 1
                if sku_candidates:
                    result.errors.append(
                        {
                            "identifier": identifier,
                            "error": f"SKU ambiguo: {len(sku_candidates)} productos comparten este SKU. Selecciona uno manualmente.",
                        }
                    )
                else:
                    result.errors.append({"identifier": identifier, "error": "Producto no encontrado en el backup"})
                continue
            row_status = ImportImagesJobRowStatus(
                row_index=row_index,
                identifier=identifier,
                product_id=int(product["id"]),
                product_name=str(product.get("name") or "") or None,
            )
            matched.append((row_status, row, product))

        job.rows = [rs for rs, _, _ in matched]
        job.total = len(matched)
        result.total_rows = len(matched)
        job.stage = "updating"

        # 2. Procesar uno por uno — secuencial para no saturar WooCommerce
        for index, (row_status, row, product) in enumerate(matched):
            identifier = row_status.identifier
            row_status.status = "running"
            job.processed = index
            job.progress = int(index * 100 / max(job.total, 1))
            job.current_product_name = row_status.product_name or identifier

            try:
                principal_raw = str(row.get("image_principal_id", "")).strip()
                gallery_raw = str(row.get("image_galeria_ids", "")).strip()
                names_raw = str(row.get("image_nombres", "")).strip()

                principal_ids = [i for i in _row_split_int_values(principal_raw) if i > 0]
                gallery_ids = [i for i in _row_split_int_values(gallery_raw) if i > 0]
                requested_names = _row_split_values(names_raw) if names_raw else []

                row_status.requested_principal_id = principal_ids[0] if principal_ids else None
                row_status.requested_gallery_ids = gallery_ids
                row_status.requested_names = requested_names

                resolved_names_pairs: list[tuple[int, str]] = []
                unresolved: list[str] = []

                if principal_ids or gallery_ids:
                    image_ids = list(dict.fromkeys(principal_ids + gallery_ids))
                elif requested_names:
                    resolved = media_service.resolve_media_by_names_from_index(
                        db=job_db, tenant_id=tenant_id, names=requested_names
                    )
                    image_ids = []
                    for r in resolved:
                        item = r.get("item") if r.get("matched") else None
                        if item and int(item.get("id") or 0) > 0:
                            iid = int(item["id"])
                            iname = str(item.get("filename") or item.get("title") or item.get("name") or "")
                            if iid not in image_ids:
                                image_ids.append(iid)
                                resolved_names_pairs.append((iid, iname))
                        else:
                            unresolved.append(str(r.get("requested") or ""))
                else:
                    image_ids = []

                row_status.resolved_image_ids = image_ids
                row_status.resolved_image_names = [name for _, name in resolved_names_pairs]
                row_status.unresolved_names = unresolved

                if not image_ids:
                    row_status.status = "skipped"
                    if not principal_raw and not gallery_raw and not names_raw:
                        row_status.error = "Sin datos de imagen en esta fila"
                    elif names_raw:
                        row_status.error = f"No se encontró en el índice: {', '.join(unresolved[:3])}{'...' if len(unresolved) > 3 else ''}"
                    else:
                        row_status.error = "Los IDs de imagen no son válidos (¿todos cero?)"
                    result.skipped += 1
                    continue

                payload = {"id": product["id"], "images": [{"id": img_id} for img_id in image_ids]}
                updated_product = await woo_service.update_product(int(product["id"]), payload)

                # Verificar que WC realmente asignó los IDs pedidos.
                applied_ids = [
                    int(img.get("id"))
                    for img in (updated_product.get("images") or [])
                    if str(img.get("id") or "").isdigit()
                ]
                row_status.applied_image_ids = applied_ids
                missing = [i for i in image_ids if i not in applied_ids]

                if missing:
                    row_status.status = "failed"
                    row_status.error = f"WooCommerce no aplicó estas imágenes: {', '.join(str(m) for m in missing)}"
                    result.failed += 1
                    result.errors.append({"identifier": identifier, "error": row_status.error})
                else:
                    row_status.status = "completed"
                    if unresolved:
                        # Aplicó lo que pudo pero algunos nombres no resolvieron — avisar.
                        row_status.error = f"No se encontraron: {', '.join(unresolved[:3])}{'...' if len(unresolved) > 3 else ''}"
                    result.updated += 1
                    updated_products.append(updated_product)

            except Exception as exc:
                row_status.status = "failed"
                row_status.error = str(exc)
                result.failed += 1
                result.errors.append({"identifier": identifier, "error": str(exc)})

        job.processed = job.total

        if updated_products:
            backup_record = backup_repository.get_backup_for_tenant(job_db, tenant_id, backup_id)
            if backup_record is not None:
                backup_service.merge_products_into_backup(job_db, backup_record, updated_products)
                job_db.commit()

        job.result = result
        job.status = "completed"
        job.stage = "completed"
        job.progress = 100
        job.current_product_name = None

    except Exception as exc:
        job.status = "failed"
        job.stage = "failed"
        job.error = str(exc)
    finally:
        if job_db is not None:
            job_db.close()


async def get_import_apply_job_status(*, job_id: str):
    job = import_apply_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import job not found")
    return job.model_dump()


async def apply_import_draft(
    *,
    db: Session,
    tenant_id: UUID,
    selected_fields: list[str],
    selected_row_indexes: list[int] | None,
    products: list[dict],
    backup_record,
    woo_service,
    credentials: dict[str, str],
    product_id_overrides: dict[int, int] | None = None,
    wp_url: str | None = None,
    wp_user: str | None = None,
    wp_app_password: str | None = None,
):
    _, file_bytes = get_current_import_draft_bytes(db, tenant_id)
    headers, rows = excel_service.parse_excel(file_bytes)
    excel_service.validate_selected_fields(headers, selected_fields)

    # Patch rows where user resolved an ambiguous SKU manually
    if product_id_overrides:
        rows = [
            {**row, "product_id": str(product_id_overrides[i]), "sku": ""} if i in product_id_overrides else row
            for i, row in enumerate(rows)
        ]

    non_image_fields = [field for field in selected_fields if field != "images"]
    image_fields = ["images"] if "images" in selected_fields else []

    response_summary = ImportApplyResult(total_rows=len(selected_row_indexes or rows))
    updated_products_for_backup: list[dict] = []

    if non_image_fields:
        summary, updated_products = await excel_service.build_bulk_update_summary(
            rows=rows,
            selected_fields=non_image_fields,
            products=products,
            woo_service=woo_service,
            selected_row_indexes=selected_row_indexes,
        )
        response_summary.updated += summary.updated
        response_summary.failed += summary.failed
        response_summary.errors.extend(summary.errors)
        response_summary.total_rows = max(response_summary.total_rows, summary.total_rows)
        updated_products_for_backup.extend(updated_products)

    if updated_products_for_backup:
        from services import backup as backup_service

        backup_service.merge_products_into_backup(db, backup_record, updated_products_for_backup)
        db.commit()

    if not image_fields:
        return response_summary.model_dump()

    media_service.require_fresh_media_index(db, backup_record.tenant_id)

    job_id = str(uuid.uuid4())
    import_apply_jobs[job_id] = ImportImagesJobStatus(
        job_id=job_id,
        status="pending",
        current_identifier="En cola",
        current_product_name="Esperando inicio del proceso de imágenes",
    )
    asyncio.create_task(
        _run_import_images_job(
            job_id=job_id,
            tenant_id=backup_record.tenant_id,
            backup_id=backup_record.id,
            rows=rows,
            selected_row_indexes=selected_row_indexes,
            products=products,
            credentials=credentials,
        )
    )
    return {
        **response_summary.model_dump(),
        "images_job": {"job_id": job_id},
    }


def delete_import_draft(db: Session, tenant_id: UUID) -> None:
    record = import_draft_repository.get_import_draft_for_tenant(db, tenant_id)
    if record is None:
        return
    _remove_file(record.storage_path)
    import_draft_repository.delete_import_draft(db, record)
    db.commit()
